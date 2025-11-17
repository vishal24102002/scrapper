import os
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import requests
from googleapiclient.discovery import build
from bs4 import BeautifulSoup

# Function to save trends to an Excel file
def save_to_excel(google_trends, youtube_trends):
    # Define file path on Desktop
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    filename = os.path.join(desktop_path, "Trending_Topics_From_Google_And_Youtube.xlsx")
    
    # Check if the file exists
    if os.path.exists(filename):
        workbook = load_workbook(filename)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Trends"
        # Add headers
        sheet.append(["Google Trends", "YouTube Trends"])
    
    # Find the last row with content
    last_row = sheet.max_row
    start_row = last_row + 3  # Leave 2 blank rows before appending
    
    # Add current date as a section header
    current_date = datetime.now().strftime("%Y-%m-%d")
    sheet.cell(row=start_row, column=1, value=f"Date: {current_date}")
    sheet.cell(row=start_row, column=2, value=f"Date: {current_date}")
    
    # Add trends to the sheet
    max_trends = max(len(google_trends), len(youtube_trends))
    for i in range(max_trends):
        google = google_trends[i] if i < len(google_trends) else ""
        youtube = youtube_trends[i] if i < len(youtube_trends) else ""
        sheet.cell(row=start_row + i + 1, column=1, value=google)
        sheet.cell(row=start_row + i + 1, column=2, value=youtube)
    
    # Save the workbook
    workbook.save(filename)
    print(f"Trends updated and saved to {filename}")

# Function to fetch Google Trends data
def fetch_google_trends(region_code='US'):
    try:
        # Calculate the date for the last 24 hours
        date = (datetime.utcnow() - timedelta(days=1)).strftime("%Y-%m-%d")
        print(f"Fetching Google Trends (Worldwide) for the last 24 hours ({date})...")

        # Construct the URL for the Google Trends daily trending searches
        trends_url = f"https://trends.google.com/trends/trendingsearches/daily/rss?geo={region_code}"
        response = requests.get(trends_url)
        response.raise_for_status()  # Raise an error for bad status codes
        
        # Parse the RSS feed data
        soup = BeautifulSoup(response.content, 'xml')
        items = soup.find_all('item')

        # Extract data
        trending_topics = []
        for item in items:
            title = item.title.text
            pub_date = item.pubDate.text

            # Handle time zone using a simpler date format
            pub_date = datetime.strptime(pub_date[:-6], "%a, %d %b %Y %H:%M:%S").strftime("%Y-%m-%d")
            
            # Check if the date matches the requested date
            if pub_date == date:
                trending_topics.append(title)

        if trending_topics:
            print("\nGoogle Trends:")
            for idx, topic in enumerate(trending_topics, start=1):
                print(f"{idx}. {topic}")
            return trending_topics
        else:
            print("No Google trends found for the last 24 hours.")
            return []
    except requests.exceptions.RequestException as e:
        print(f"Error fetching Google trends: {e}")
        return []
    except ValueError as e:
        print(f"Error parsing date: {e}")
        return []

# Function to fetch trending videos on YouTube
def fetch_trending_videos(api_key, region_code='US'):
    try:
        youtube = build('youtube', 'v3', developerKey=api_key)
        print(f"Fetching YouTube trending videos for region: {region_code}...")

        request = youtube.videos().list(
            part="snippet",
            chart="mostPopular",
            regionCode=region_code,
            maxResults=10
        )
        response = request.execute()

        trending_videos = []
        print("\nYouTube Trends:")
        for idx, video in enumerate(response.get('items', []), start=1):
            title = video['snippet']['title']
            hashtags = [tag for tag in video['snippet'].get('tags', [])] if 'tags' in video['snippet'] else []
            print(f"{idx}. Title: {title}")
            print(f"   Hashtags: {', '.join(hashtags) if hashtags else 'No hashtags'}")
            trending_videos.append(title)

        return trending_videos
    except Exception as e:
        print(f"Error fetching YouTube trending videos: {e}")
        return []

# Main function to execute
if __name__ == "__main__":
    # API Key for YouTube Data API
    API_KEY = "AIzaSyCGfb_-oeu9UXNxeAu9L3nYeRIeni2O14s"  # Replace with your API key

    # Fetch Google Trends
    google_trends = fetch_google_trends(region_code='US')
    if not google_trends:
        print("No significant Google Trends found.")

    # Fetch YouTube Trends
    youtube_trends = fetch_trending_videos(API_KEY, region_code="US")
    if not youtube_trends:
        print("No significant YouTube trends found.")

    # Save trends to Excel
    save_to_excel(google_trends, youtube_trends)
